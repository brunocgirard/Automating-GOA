import html
import re
import os
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Any

from openpyxl import load_workbook

# Define constants
TEMPLATE_DIR = Path("templates")
EXCEL_FILENAME = "GOA_template.xlsx"
EXCEL_PATH = TEMPLATE_DIR / EXCEL_FILENAME
OUTPUT_HTML_FILENAME = "goa_form.html"
OUTPUT_HTML_PATH = TEMPLATE_DIR / OUTPUT_HTML_FILENAME

def load_rows(excel_path: Path = EXCEL_PATH) -> List[Dict[str, str]]:
    """
    Reads rows from the Excel template.
    Returns a list of dictionaries representing each field.
    """
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel template not found at {excel_path}")

    wb = load_workbook(excel_path, data_only=True)
    if "Form" not in wb.sheetnames:
        raise ValueError(f"Sheet 'Form' not found in {excel_path}")
        
    ws = wb["Form"]
    rows = []
    
    # Iterate rows, skipping header
    for raw in ws.iter_rows(min_row=2, values_only=True):
        # Ensure we have enough columns (at least 6)
        if len(raw) < 6:
            continue
            
        section, subsection, subsub, field_name, ftype, placeholder = raw[:6]
        
        # Skip empty rows or rows without critical info
        if not placeholder or not field_name:
            continue
            
        rows.append(
            {
                "section": str(section or "").strip(),
                "subsection": str(subsection or "").strip(),
                "subsub": str(subsub or "").strip(),
                "field": str(field_name or "").strip(),
                "type": str(ftype or "").strip().lower(),
                "placeholder": str(placeholder).strip(),
            }
        )

    # Ensure we always have an options_listing placeholder for downstream rendering.
    # If it's missing, first try to reuse the existing "Option Listing" row (e.g., f0091)
    # by renaming its placeholder to options_listing; if not found, insert a new row
    # immediately after the Option Listing section to preserve the Excel order.
    has_options_listing = any(r["placeholder"] == "options_listing" for r in rows)
    if not has_options_listing:
        renamed = False
        for r in rows:
            sec = r["section"].strip().lower()
            fld = r["field"].strip().lower()
            if sec == "option listing" or fld.startswith("option listing"):
                r["placeholder"] = "options_listing"
                r["type"] = "textarea"
                renamed = True
                break

        if not renamed:
            insert_at = len(rows)
            for idx, r in enumerate(rows):
                if r["section"].strip().lower() == "option listing":
                    insert_at = idx + 1
                    break

            rows.insert(
                insert_at,
                {
                    "section": "Option Listing",
                    "subsection": "",
                    "subsub": "",
                    "field": "Options Listing (auto-generated)",
                    "type": "textarea",
                    "placeholder": "options_listing",
                },
            )
    return rows

def display_label(raw: str) -> str:
    """Strip helper suffixes like (text)/(checkbox)/(qty) and example hints."""
    label = re.sub(r"\s*-\s*example:.*", "", raw, flags=re.IGNORECASE)
    label = re.sub(r"\s*\((text|checkbox|qty|text heading|heading)[^)]*\)", "", label, flags=re.IGNORECASE)
    return label.strip(" -:")

def group_by_section(rows):
    sections = defaultdict(list)
    for row in rows:
        sections[row["section"]].append(row)
    return sections

def render_input(row: dict) -> str:
    ph = html.escape(f"{{{{{row['placeholder']}}}}}")
    label = html.escape(display_label(row["field"]))
    
    if row["type"] == "checkbox":
        return f"""
        <label class="field checkbox" data-placeholder="{ph}">
          <input type="checkbox" name="{row['placeholder']}" data-placeholder="{ph}" />
          <span class="label">{label}</span>
          <span class="token">{ph}</span>
        </label>
        """
    
    # Special handling for options_listing or explicit textarea type
    if row['placeholder'] == 'options_listing' or row["type"] == "textarea":
        return f"""
        <label class="field textarea" data-placeholder="{ph}">
          <span class="label">{label}</span>
          <textarea name="{row['placeholder']}" data-placeholder="{ph}" rows="5"></textarea>
          <span class="token">{ph}</span>
        </label>
        """

    input_type = "number" if row["type"] == "qty" else "text"
    return f"""
    <label class="field" data-placeholder="{ph}">
      <span class="label">{label}</span>
      <input type="{input_type}" name="{row['placeholder']}" data-placeholder="{ph}" />
      <span class="token">{ph}</span>
    </label>
    """

def render_group(title: str, items: list[dict]) -> str:
    all_check = all(it["type"] == "checkbox" for it in items)
    grid_class = "checkbox-grid" if all_check else "field-grid"
    heading = f'<div class="group-title">{html.escape(title)}</div>' if title else ""
    fields_html = "\n".join(render_input(it) for it in items)
    return f"""
    <div class="group">
      {heading}
      <div class="{grid_class}">
        {fields_html}
      </div>
    </div>
    """

def render_section(name: str, items: list[dict]) -> str:
    clean_name = re.sub(r"\s*\(section\)", "", name, flags=re.IGNORECASE).strip()
    # bucket by subsection/subsub
    grouped = []
    current = (None, None)
    bucket = []
    for entry in items:
        key = (entry["subsection"], entry["subsub"])
        if key != current and bucket:
            grouped.append((current, bucket))
            bucket = []
        current = key
        bucket.append(entry)
    if bucket:
        grouped.append((current, bucket))

    groups_html = ""
    for (sub, subsub), bucket in grouped:
        title_parts = [p for p in (sub, subsub) if p]
        title = " / ".join(title_parts)
        groups_html += render_group(title, bucket)

    return f"""
    <section class="section">
      <div class="section-header">
        <h2>{html.escape(clean_name)}</h2>
      </div>
      {groups_html}
    </section>
    """

def build_html(rows):
    sections = group_by_section(rows)
    # Sort sections to ensure consistent order if needed, or rely on Excel order
    # Here we rely on Excel order preserved in 'rows' list, but group_by_section uses defaultdict which might lose order if python < 3.7 (unlikely)
    # Better to iterate sections in order of appearance
    section_order = []
    seen = set()
    for row in rows:
        if row["section"] not in seen:
            section_order.append(row["section"])
            seen.add(row["section"])
            
    body = "\n".join(render_section(name, sections[name]) for name in section_order)
    
    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width,initial-scale=1" />
  <title>General Order Acknowledgement</title>
  <style>
    :root {{
      --bg: #f3f5f8;
      --card: #ffffff;
      --ink: #1f2430;
      --muted: #4b5563;
      --accent: #c00000;      /* template docx red accent */
      --header-fill: #e5e5e5; /* template docx gray strip */
      --border: #cdd4e0;
    }}
    * {{ box-sizing: border-box; font-family: "Calibri", "Segoe UI", Arial, sans-serif; }}
    body {{
      margin: 0;
      padding: 24px;
      background: var(--bg);
      color: var(--ink);
    }}
    .page {{ max-width: 1200px; margin: 0 auto; }}
    header {{ margin-bottom: 18px; display: flex; justify-content: space-between; align-items: center; }}
    h1 {{
      margin: 0 0 6px;
      font-size: 26px;
      color: var(--accent);
      font-weight: 700;
      letter-spacing: -0.015em;
    }}
    .subtitle {{ color: var(--muted); margin: 0 0 8px; font-size: 14px; }}
    .note {{ font-size: 13px; color: var(--muted); margin: 3px 0; }}
    .divider {{
      height: 6px;
      background: var(--header-fill);
      border: 1px solid var(--border);
      border-radius: 6px;
      margin-bottom: 12px;
    }}
    .section {{
      background: var(--card);
      border: 1px solid var(--border);
      border-radius: 10px;
      margin-bottom: 14px;
      box-shadow: 0 4px 12px rgba(0,0,0,0.04);
      overflow: hidden;
    }}
    .section-header {{
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 10px;
      padding: 10px 14px;
      background: var(--header-fill);
      border-bottom: 1px solid var(--border);
      border-top-left-radius: 10px;
      border-top-right-radius: 10px;
      cursor: pointer;
      user-select: none;
    }}
    .section-header.active + .section-content {{
        /* styles when open */
    }}
    .section.collapsed > .section-header {{
        border-bottom-color: transparent;
    }}
    .toggle-icon {{
        transition: transform 0.3s ease;
        font-weight: bold;
        font-size: 20px;
        color: var(--muted);
    }}
    .section-header.active .toggle-icon {{
        transform: rotate(45deg);
    }}
    .section-content {{
        padding: 16px 14px;
        overflow: hidden;
        max-height: 10000px; /* A large enough value to not clip content */
        transition: max-height 0.4s ease-in-out, padding 0.3s ease-in-out;
    }}
    .section.collapsed > .section-content {{
        max-height: 0;
        padding-top: 0;
        padding-bottom: 0;
    }}
    .section h2 {{
      margin: 0;
      font-size: 18px;
      color: var(--accent);
      font-weight: 700;
    }}
    .pill {{
      background: var(--header-fill);
      color: var(--accent);
      padding: 4px 10px;
      border-radius: 999px;
      font-size: 11px;
      font-weight: 700;
      border: 1px solid var(--border);
    }}
    .group {{ margin: 10px 14px 0 14px; }}
    .group-title {{
      font-size: 13px;
      font-weight: 700;
      color: var(--muted);
      margin: 6px 0 4px;
    }}
    .field-grid {{
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(260px, 1fr));
      gap: 10px 12px;
    }}
    .checkbox-grid {{
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
      gap: 8px 10px;
    }}
    .field {{
      border-left: 3px solid #f3f3f3;
    }}
    .field {{
      display: grid;
      grid-template-rows: auto auto auto;
      gap: 4px;
      padding: 10px;
      border: 1px solid var(--border);
      border-radius: 6px;
      background: #fff;
    }}
    .field.checkbox {{
      grid-template-columns: auto 1fr auto;
      grid-template-rows: auto;
      align-items: center;
      gap: 8px;
      background: #f8f9fc;
      border-color: #d5dbe7;
    }}
    .field.checkbox .label {{ font-weight: 600; color: var(--ink); }}
    .label {{ font-size: 13px; color: var(--ink); font-weight: 600; }}
    input[type="text"], input[type="number"], textarea {{
      width: 100%;
      padding: 7px 9px;
      border-radius: 4px;
      border: 1px solid var(--border);
      background: #fdfdff;
      font-size: 14px;
      color: var(--ink);
      font-family: inherit;
    }}
    textarea {{ resize: vertical; }}
    input:focus {{ outline: 2px solid #b5c7e3; }}
    .token {{
      font-size: 12px;
      color: var(--muted);
      font-family: "Consolas", "SFMono-Regular", monospace;
      display: none; /* hide placeholders */
    }}
    .header-fill {{
      background: var(--header-fill);
      border: 1px solid var(--border);
      border-radius: 6px;
      padding: 6px 10px;
      font-weight: 700;
      color: var(--accent);
      display: inline-block;
      margin-top: 6px;
    }}
    input[type="checkbox"] {{
      width: 18px;
      height: 18px;
    }}
    
    /* Styles for Read-Only / Filled View */
    body.readonly input {{
        border: none;
        background: transparent;
        pointer-events: none;
    }}
    body.readonly input[type="checkbox"] {{
        /* Custom styling for checked box in print mode? Or just keep browser default */
    }}
    body.readonly .field {{
        border: 1px solid transparent; /* Hide border or make lighter */
        box-shadow: none;
        background: transparent;
    }}
    body.readonly .section {{
        box-shadow: none;
        border: 1px solid #eee;
    }}

    /* Edit Mode Styles */
    .delete-btn {{
        display: none;
        padding: 4px 8px;
        background: #dc2626;
        color: white;
        border: none;
        border-radius: 4px;
        cursor: pointer;
        font-size: 12px;
        font-weight: bold;
        opacity: 0;
        transition: opacity 0.2s;
    }}
    body.edit-mode .delete-btn {{
        display: inline-block;
    }}
    body.edit-mode .section:hover .delete-section-btn,
    body.edit-mode .field:hover .delete-field-btn {{
        opacity: 1;
    }}
    body.edit-mode .section-header h2[contenteditable="true"],
    body.edit-mode .field .label[contenteditable="true"],
    body.edit-mode .group-title[contenteditable="true"] {{
        cursor: text;
        padding: 4px 6px;
        border-radius: 4px;
        transition: background 0.2s, outline 0.2s;
        display: inline-block;
        min-width: 50px;
    }}
    body.edit-mode .section-header h2[contenteditable="true"]:hover,
    body.edit-mode .field .label[contenteditable="true"]:hover,
    body.edit-mode .group-title[contenteditable="true"]:hover {{
        background: rgba(37, 99, 235, 0.15);
        outline: 2px dashed #2563eb;
    }}
    body.edit-mode .section-header h2[contenteditable="true"]:focus,
    body.edit-mode .field .label[contenteditable="true"]:focus,
    body.edit-mode .group-title[contenteditable="true"]:focus {{
        background: rgba(37, 99, 235, 0.25);
        outline: 2px solid #2563eb;
    }}
    /* Add visual indicator for editable group titles */
    body.edit-mode .group-title[contenteditable="true"]::before {{
        content: '✎ ';
        color: #2563eb;
        font-weight: bold;
        margin-right: 4px;
        opacity: 0.6;
    }}
    .delete-section-btn {{
        margin-left: auto;
    }}
    .delete-field-btn {{
        position: absolute;
        top: 4px;
        right: 4px;
    }}
    body.edit-mode .field {{
        position: relative;
    }}
    body.edit-mode .section {{
        border: 2px dashed transparent;
        transition: border-color 0.2s;
    }}
    body.edit-mode .section:hover {{
        border-color: #cbd5e1;
    }}

    /* Section Controls Styling */
    .section-controls button {{
        transition: all 0.2s ease;
    }}
    .section-controls button:hover {{
        opacity: 0.9;
        transform: translateY(-1px);
        box-shadow: 0 2px 8px rgba(0,0,0,0.15);
    }}
    .section-controls button:active {{
        transform: translateY(0);
        box-shadow: 0 1px 4px rgba(0,0,0,0.1);
    }}

    /* Print Styles - Enhanced for PDF Export */
    @page {{
        size: letter portrait;
        margin: 10mm;
    }}

    @media print {{
        .no-print {{ display: none !important; }}
        body {{
            background: white;
            padding: 0;
            margin: 0;
            -webkit-print-color-adjust: exact;
            print-color-adjust: exact;
            color-adjust: exact;
        }}
        .page {{
            max-width: 100%;
            margin: 0;
            width: 100%;
        }}
        .section {{
            border: 1px solid #ccc;
            page-break-inside: auto;
            margin-bottom: 10px;
            box-shadow: none;
        }}
        .section-header {{
            page-break-after: avoid;
            break-after: avoid;
            page-break-inside: avoid;
            background: #e5e5e5 !important;
            -webkit-print-color-adjust: exact;
            print-color-adjust: exact;
        }}
        /* Ensure sections are visible (not collapsed) */
        .section.collapsed > .section-content {{
            max-height: none !important;
            padding: 16px 14px !important;
        }}
        .section-header .toggle-icon {{
            display: none;
        }}
        .group {{
            page-break-inside: auto;
            margin-bottom: 8px;
        }}
        .group-title {{
            page-break-after: avoid;
            break-after: avoid;
            font-weight: bold;
        }}
        .field {{
            border: none;
            break-inside: avoid;
            page-break-inside: avoid;
            margin-bottom: 4px;
        }}
        .field-grid, .checkbox-grid {{
            page-break-inside: auto;
        }}
        input, textarea {{
            background: transparent;
            border: none;
            border-bottom: 1px solid #ddd;
            color: #000;
        }}
        /* Prevent orphaned headers */
        h1, h2, .section-header, .group-title {{
            orphans: 3;
            widows: 3;
        }}
        /* Ensure proper spacing */
        .divider {{
            background: #e5e5e5 !important;
            -webkit-print-color-adjust: exact;
            print-color-adjust: exact;
        }}
    }}

    @media (max-width: 640px) {{
      body {{ padding: 12px; }}
      .section {{ padding: 12px; }}
      .field-grid,
      .checkbox-grid {{
        grid-template-columns: 1fr;
      }}
      input[type="text"], input[type="number"] {{
        padding: 10px 12px;
        font-size: 15px;
      }}
      .field {{
        gap: 6px;
      }}
      .field.checkbox {{
        grid-template-columns: auto 1fr;
        grid-template-rows: auto auto;
        row-gap: 4px;
      }}
    }}
  </style>
</head>
  <body>
    <div class="page">
      <header>
        <h1>General Order Acknowledgement</h1>
        <div class="no-print" style="display: flex; gap: 10px;">
          <button id="editModeBtn" onclick="toggleEditMode()" style="padding: 8px 16px; background: #2563eb; color: white; border: none; border-radius: 4px; cursor: pointer; font-weight: bold;">Enable Edit Mode</button>
          <button id="downloadBtn" onclick="downloadModifiedHTML()" style="padding: 8px 16px; background: #059669; color: white; border: none; border-radius: 4px; cursor: pointer; font-weight: bold; display: none;">Download Modified Form</button>
          <button onclick="saveFilledFormHTML()" style="padding: 8px 16px; background: #0891b2; color: white; border: none; border-radius: 4px; cursor: pointer; font-weight: bold;">Save Filled Form (HTML)</button>
          <button onclick="generatePDF()" style="padding: 8px 16px; background: #c00000; color: white; border: none; border-radius: 4px; cursor: pointer; font-weight: bold;">Print to PDF</button>
        </div>
      </header>
      <div class="divider"></div>

      <!-- Section Controls -->
      <div class="section-controls no-print" style="margin-bottom: 14px; display: flex; gap: 10px; justify-content: flex-end;">
        <button onclick="expandAllSections()" style="padding: 6px 14px; background: #10b981; color: white; border: none; border-radius: 4px; cursor: pointer; font-weight: 600; font-size: 14px;">Expand All Sections</button>
        <button onclick="collapseAllSections()" style="padding: 6px 14px; background: #6b7280; color: white; border: none; border-radius: 4px; cursor: pointer; font-weight: 600; font-size: 14px;">Collapse All Sections</button>
      </div>

      {body}
    </div>
    <script>
      document.addEventListener('DOMContentLoaded', function () {{
          const sections = document.querySelectorAll('.section');
      
          sections.forEach((section, index) => {{
              const header = section.querySelector('.section-header');
              if (header) {{
                  // Create content wrapper
                  const content = document.createElement('div');
                  content.className = 'section-content';
      
                  // Move all group elements into the content wrapper
                  const groups = Array.from(section.children).filter(child => child.classList.contains('group'));
                  groups.forEach(group => {{
                      content.appendChild(group);
                  }});
                  section.appendChild(content);

                  // Add toggle icon only if it doesn't already exist
                  if (!header.querySelector('.toggle-icon')) {{
                      const icon = document.createElement('span');
                      icon.className = 'toggle-icon';
                      icon.textContent = '+';
                      header.appendChild(icon);
                  }}

                  // Add click listener
                  header.addEventListener('click', () => {{
                      section.classList.toggle('collapsed');
                      header.classList.toggle('active');
                  }});
      
                  // Initially collapse all but the first section
                  if (index > 0) {{
                      section.classList.add('collapsed');
                  }} else {{
                      header.classList.add('active');
                  }}
              }}
          }});

          // Add delete buttons and make elements editable for edit mode
          initializeEditMode();
      }});

      // Initialize edit mode features
      function initializeEditMode() {{
          const sections = document.querySelectorAll('.section');

          sections.forEach(section => {{
              const header = section.querySelector('.section-header');
              const h2 = header.querySelector('h2');

              // Add delete button for section
              const deleteSecBtn = document.createElement('button');
              deleteSecBtn.className = 'delete-btn delete-section-btn';
              deleteSecBtn.textContent = '× Delete Section';
              deleteSecBtn.onclick = (e) => {{
                  e.stopPropagation();
                  deleteSection(section);
              }};
              header.appendChild(deleteSecBtn);

              // Add delete buttons for fields
              const fields = section.querySelectorAll('.field');
              fields.forEach(field => {{
                  const deleteFieldBtn = document.createElement('button');
                  deleteFieldBtn.className = 'delete-btn delete-field-btn';
                  deleteFieldBtn.textContent = '×';
                  deleteFieldBtn.onclick = (e) => {{
                      e.stopPropagation();
                      deleteField(field);
                  }};
                  field.appendChild(deleteFieldBtn);
              }});
          }});
      }}

      // Toggle edit mode
      function toggleEditMode() {{
          const body = document.body;
          const editBtn = document.getElementById('editModeBtn');
          const downloadBtn = document.getElementById('downloadBtn');

          body.classList.toggle('edit-mode');
          const isEditMode = body.classList.contains('edit-mode');

          if (isEditMode) {{
              editBtn.textContent = 'Disable Edit Mode';
              editBtn.style.background = '#dc2626';
              downloadBtn.style.display = 'inline-block';
              enableEditing();
          }} else {{
              editBtn.textContent = 'Enable Edit Mode';
              editBtn.style.background = '#2563eb';
              downloadBtn.style.display = 'none';
              disableEditing();
          }}
      }}

      // Enable editing
      function enableEditing() {{
          // Make section headers editable
          document.querySelectorAll('.section-header h2').forEach(h2 => {{
              h2.setAttribute('contenteditable', 'true');
              h2.setAttribute('title', 'Click to edit section name');
          }});

          // Make subsection/group titles editable
          document.querySelectorAll('.group-title').forEach(groupTitle => {{
              groupTitle.setAttribute('contenteditable', 'true');
              groupTitle.setAttribute('title', 'Click to edit subsection name');
          }});

          // Make field labels editable
          document.querySelectorAll('.field .label').forEach(label => {{
              label.setAttribute('contenteditable', 'true');
              label.setAttribute('title', 'Click to edit field label');
          }});
      }}

      // Disable editing
      function disableEditing() {{
          document.querySelectorAll('[contenteditable="true"]').forEach(el => {{
              el.removeAttribute('contenteditable');
              el.removeAttribute('title');
          }});
      }}

      // Delete a field
      function deleteField(field) {{
          if (confirm('Are you sure you want to delete this field?')) {{
              field.remove();
          }}
      }}

      // Delete a section
      function deleteSection(section) {{
          const sectionName = section.querySelector('h2').textContent;
          if (confirm(`Are you sure you want to delete the entire "${{sectionName}}" section?`)) {{
              section.remove();
          }}
      }}

      // Download modified HTML
      function downloadModifiedHTML() {{
          // Clone the document
          const clone = document.documentElement.cloneNode(true);

          // Remove edit mode class from body
          const cloneBody = clone.querySelector('body');
          cloneBody.classList.remove('edit-mode');

          // Remove contenteditable attributes
          clone.querySelectorAll('[contenteditable="true"]').forEach(el => {{
              el.removeAttribute('contenteditable');
              el.removeAttribute('title');
          }});

          // Remove delete buttons
          clone.querySelectorAll('.delete-btn').forEach(btn => btn.remove());

          // Generate HTML string
          const htmlString = '<!DOCTYPE html>\\n' + clone.outerHTML;

          // Create blob and download
          const blob = new Blob([htmlString], {{ type: 'text/html' }});
          const url = URL.createObjectURL(blob);
          const a = document.createElement('a');
          a.href = url;
          a.download = 'goa_form_modified.html';
          document.body.appendChild(a);
          a.click();
          document.body.removeChild(a);
          URL.revokeObjectURL(url);

          alert('Modified form downloaded! You can now use this customized form.');
      }}

      // Save filled form as HTML with all data preserved
      function saveFilledFormHTML() {{
          // Clone the document
          const clone = document.documentElement.cloneNode(true);

          // Remove edit mode class from body
          const cloneBody = clone.querySelector('body');
          cloneBody.classList.remove('edit-mode');

          // Remove contenteditable attributes
          clone.querySelectorAll('[contenteditable="true"]').forEach(el => {{
              el.removeAttribute('contenteditable');
              el.removeAttribute('title');
          }});

          // Remove delete buttons
          clone.querySelectorAll('.delete-btn').forEach(btn => btn.remove());

          // Preserve all input values
          document.querySelectorAll('input[type="text"], input[type="number"]').forEach((input, index) => {{
              const cloneInputs = clone.querySelectorAll('input[type="text"], input[type="number"]');
              if (cloneInputs[index] && input.value) {{
                  cloneInputs[index].setAttribute('value', input.value);
              }}
          }});

          // Preserve all checkbox states
          document.querySelectorAll('input[type="checkbox"]').forEach((checkbox, index) => {{
              const cloneCheckboxes = clone.querySelectorAll('input[type="checkbox"]');
              if (cloneCheckboxes[index]) {{
                  if (checkbox.checked) {{
                      cloneCheckboxes[index].setAttribute('checked', 'checked');
                  }} else {{
                      cloneCheckboxes[index].removeAttribute('checked');
                  }}
              }}
          }});

          // Preserve all textarea values
          document.querySelectorAll('textarea').forEach((textarea, index) => {{
              const cloneTextareas = clone.querySelectorAll('textarea');
              if (cloneTextareas[index] && textarea.value) {{
                  cloneTextareas[index].textContent = textarea.value;
              }}
          }});

          // Generate HTML string
          const htmlString = '<!DOCTYPE html>\\n' + clone.outerHTML;

          // Create blob and download
          const blob = new Blob([htmlString], {{ type: 'text/html' }});
          const url = URL.createObjectURL(blob);
          const a = document.createElement('a');
          a.href = url;
          a.download = 'goa_form_filled.html';
          document.body.appendChild(a);
          a.click();
          document.body.removeChild(a);
          URL.revokeObjectURL(url);

          alert('Filled form saved as HTML! You can reopen this file to make revisions and re-save as PDF.');
      }}

      // Generate PDF using browser's native print-to-PDF functionality
      function generatePDF() {{
          // Store current edit mode state
          const wasInEditMode = document.body.classList.contains('edit-mode');

          // Temporarily disable edit mode for PDF generation
          if (wasInEditMode) {{
              document.body.classList.remove('edit-mode');
          }}

          // Store collapsed sections state and expand all sections for PDF
          const sections = document.querySelectorAll('.section');
          const collapsedSections = [];

          sections.forEach((section, index) => {{
              if (section.classList.contains('collapsed')) {{
                  collapsedSections.push(index);
                  section.classList.remove('collapsed');
                  const header = section.querySelector('.section-header');
                  if (header) {{
                      header.classList.add('active');
                  }}
              }}
          }});

          // Wait for DOM to fully render expanded sections before printing
          setTimeout(() => {{
              // Trigger browser's print dialog (user can save as PDF)
              window.print();

              // Restore collapsed sections after print dialog is handled
              // Note: This happens immediately, but browser waits for print dialog to close
              setTimeout(() => {{
                  collapsedSections.forEach(index => {{
                      sections[index].classList.add('collapsed');
                      const header = sections[index].querySelector('.section-header');
                      if (header) {{
                          header.classList.remove('active');
                      }}
                  }});

                  // Restore edit mode if it was active
                  if (wasInEditMode) {{
                      document.body.classList.add('edit-mode');
                  }}
              }}, 100);
          }}, 300); // Brief delay to ensure sections are fully expanded and rendered
      }}

      // Expand all sections
      function expandAllSections() {{
          document.querySelectorAll('.section').forEach(section => {{
              section.classList.remove('collapsed');
              const header = section.querySelector('.section-header');
              if (header) {{
                  header.classList.add('active');
              }}
          }});
      }}

      // Collapse all sections
      function collapseAllSections() {{
          document.querySelectorAll('.section').forEach(section => {{
              section.classList.add('collapsed');
              const header = section.querySelector('.section-header');
              if (header) {{
                  header.classList.remove('active');
              }}
          }});
      }}
    </script>
  </body>
</html>
"""

def generate_goa_form(excel_path: Path = EXCEL_PATH, output_path: Path = OUTPUT_HTML_PATH) -> bool:
    """
    Generates the HTML form from the Excel template.
    """
    try:
        print(f"Generating form from {excel_path}...")
        rows = load_rows(excel_path)
        html_doc = build_html(rows)
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(html_doc)
        print(f"Successfully generated {output_path} with {len(rows)} fields.")
        return True
    except Exception as e:
        print(f"Error generating form: {e}")
        import traceback
        traceback.print_exc()
        return False

def extract_schema_from_excel(excel_path: Path = EXCEL_PATH) -> Dict[str, Dict]:
    """
    Extracts a schema dictionary from the Excel template for use by the LLM.
    
    Returns:
    {
        "placeholder_key": {
            "type": "string" | "boolean",
            "section": "Section Name",
            "subsection": "Subsection Name",
            "description": "Full description for LLM",
            "synonyms": [...],
            "positive_indicators": [...]
        }
    }
    """
    try:
        rows = load_rows(excel_path)
        schema = {}
        
        # We need to import helper functions for synonyms if we want to reuse them
        # or implement simplified versions here.
        # Let's define simple helpers here to avoid circular imports if possible,
        # or duplicate the logic from template_utils.py slightly modified.
        
        from src.utils.template_utils import generate_synonyms_for_checkbox, generate_positive_indicators
        
        for row in rows:
            ph_key = row["placeholder"]
            ftype = "boolean" if row["type"] == "checkbox" else "string"
            
            # Construct a rich description
            parts = [row["section"], row["subsection"], row["subsub"], row["field"]]
            description = " - ".join(filter(None, parts))
            
            schema[ph_key] = {
                "type": ftype,
                "section": row["section"],
                "subsection": row["subsection"],
                "description": description,
                "location": "form"
            }
            
            if ftype == "boolean":
                # Generate synonyms and indicators
                synonyms = generate_synonyms_for_checkbox(ph_key, description)
                schema[ph_key]["synonyms"] = synonyms
                schema[ph_key]["positive_indicators"] = generate_positive_indicators(ph_key, description, synonyms)
                
        return schema
        
    except Exception as e:
        print(f"Error extracting schema from Excel: {e}")
        return {}

def get_all_fields_from_excel(excel_path: Path = EXCEL_PATH) -> Dict[str, str]:
    """
    Returns a simple dictionary of {placeholder: description} for all fields.
    Used for the CRM editor "Add Field" dropdown.
    """
    try:
        rows = load_rows(excel_path)
        fields = {}
        for row in rows:
            ph_key = row["placeholder"]
            parts = [row["section"], row["subsection"], row["subsub"], row["field"]]
            description = " - ".join(filter(None, parts))
            fields[ph_key] = description
        return fields
    except Exception as e:
        print(f"Error getting fields from Excel: {e}")
        return {}

if __name__ == "__main__":
    generate_goa_form()
