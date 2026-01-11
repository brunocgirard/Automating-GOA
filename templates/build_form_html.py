import html
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

EXCEL_PATH = Path("GOA_template.xlsx")
OUTPUT_HTML = Path("goa_form.html")


def load_rows():
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Form"]
    rows = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        section, subsection, subsub, field_name, ftype, placeholder = raw[:6]
        if not placeholder or not field_name:
            continue
        rows.append(
            {
                "section": section or "",
                "subsection": subsection or "",
                "subsub": subsub or "",
                "field": (field_name or "").strip(),
                "type": (ftype or "").strip().lower(),
                "placeholder": placeholder.strip(),
            }
        )
    return rows


def display_label(raw: str) -> str:
    """Strip helper suffixes like (text)/(checkbox)/(qty) and example hints."""
    label = re.sub(r"\s*-\s*example:.*", "", raw, flags=re.IGNORECASE)
    label = re.sub(r"\s*\((text|checkbox|qty|text heading|heading)[^)]*\)", "", label, flags=re.IGNORECASE)
    return label.strip(" -:")


def group_by_section(rows):
    sections = defaultdict(list)
    for row in rows:
        sections[row["section"]].append(row)
    return sections


def render_input(row: dict) -> str:
    ph = html.escape(f"{{{{{row['placeholder']}}}}}")
    label = html.escape(display_label(row["field"]))
    if row["type"] == "checkbox":
        return f"""
        <label class="field checkbox" data-placeholder="{ph}">
          <input type="checkbox" name="{row['placeholder']}" data-placeholder="{ph}" />
          <span class="label">{label}</span>
          <span class="token">{ph}</span>
        </label>
        """
    input_type = "number" if row["type"] == "qty" else "text"
    return f"""
    <label class="field" data-placeholder="{ph}">
      <span class="label">{label}</span>
      <input type="{input_type}" name="{row['placeholder']}" data-placeholder="{ph}" />
      <span class="token">{ph}</span>
    </label>
    """


def render_group(title: str, items: list[dict]) -> str:
    all_check = all(it["type"] == "checkbox" for it in items)
    grid_class = "checkbox-grid" if all_check else "field-grid"
    heading = f'<div class="group-title">{html.escape(title)}</div>' if title else ""
    fields_html = "\n".join(render_input(it) for it in items)
    return f"""
    <div class="group">
      {heading}
      <div class="{grid_class}">
        {fields_html}
      </div>
    </div>
    """


def render_section(name: str, items: list[dict]) -> str:
    clean_name = re.sub(r"\s*\(section\)", "", name, flags=re.IGNORECASE).strip()
    # bucket by subsection/subsub
    grouped = []
    current = (None, None)
    bucket = []
    for entry in items:
        key = (entry["subsection"], entry["subsub"])
        if key != current and bucket:
            grouped.append((current, bucket))
            bucket = []
        current = key
        bucket.append(entry)
    if bucket:
        grouped.append((current, bucket))

    groups_html = ""
    for (sub, subsub), bucket in grouped:
        title_parts = [p for p in (sub, subsub) if p]
        title = " / ".join(title_parts)
        groups_html += render_group(title, bucket)

    return f"""
    <section class="section">
      <div class="section-header">
        <h2>{html.escape(clean_name)}</h2>
      </div>
      {groups_html}
    </section>
    """


def build_html(rows):
    sections = group_by_section(rows)
    body = "\n".join(render_section(name, items) for name, items in sections.items())
    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width,initial-scale=1" />
  <title>General Order Acknowledgement (HTML)</title>
  <style>
    :root {{
      --bg: #f3f5f8;
      --card: #ffffff;
      --ink: #1f2430;
      --muted: #4b5563;
      --accent: #c00000;      /* template docx red accent */
      --header-fill: #e5e5e5; /* template docx gray strip */
      --border: #cdd4e0;
    }}
    * {{ box-sizing: border-box; font-family: "Calibri", "Segoe UI", Arial, sans-serif; }}
    body {{
      margin: 0;
      padding: 24px;
      background: var(--bg);
      color: var(--ink);
    }}
    .page {{ max-width: 1200px; margin: 0 auto; }}
    header {{ margin-bottom: 18px; }}
    h1 {{
      margin: 0 0 6px;
      font-size: 26px;
      color: var(--accent);
      font-weight: 700;
      letter-spacing: -0.015em;
    }}
    .subtitle {{ color: var(--muted); margin: 0 0 8px; font-size: 14px; }}
    .note {{ font-size: 13px; color: var(--muted); margin: 3px 0; }}
    .divider {{
      height: 6px;
      background: var(--header-fill);
      border: 1px solid var(--border);
      border-radius: 6px;
      margin-bottom: 12px;
    }}
    .section {
      background: var(--card);
      border: 1px solid var(--border);
      border-radius: 10px;
      margin-bottom: 14px;
      box-shadow: 0 4px 12px rgba(0,0,0,0.04);
      overflow: hidden;
    }
    .section-header {
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 10px;
      padding: 10px 14px;
      background: var(--header-fill);
      border-bottom: 1px solid var(--border);
      border-top-left-radius: 10px;
      border-top-right-radius: 10px;
      cursor: pointer;
      user-select: none;
    }
    .section-header.active + .section-content {
        /* styles when open */
    }
    .section.collapsed > .section-header {
        border-bottom-color: transparent;
    }
    .toggle-icon {
        transition: transform 0.3s ease;
        font-weight: bold;
        font-size: 20px;
        color: var(--muted);
    }
    .section-header.active .toggle-icon {
        transform: rotate(45deg);
    }
    .section-content {
        padding: 16px 14px;
        overflow: hidden;
        max-height: 10000px; /* A large enough value to not clip content */
        transition: max-height 0.4s ease-in-out, padding 0.3s ease-in-out;
    }
    .section.collapsed > .section-content {
        max-height: 0;
        padding-top: 0;
        padding-bottom: 0;
    }
    .section h2 {{
      margin: 0;
      font-size: 18px;
      color: var(--accent);
      font-weight: 700;
    }}
    .pill {{
      background: var(--header-fill);
      color: var(--accent);
      padding: 4px 10px;
      border-radius: 999px;
      font-size: 11px;
      font-weight: 700;
      border: 1px solid var(--border);
    }}
    .group {{ margin: 10px 14px 0 14px; }}
    .group-title {{
      font-size: 13px;
      font-weight: 700;
      color: var(--muted);
      margin: 6px 0 4px;
    }}
    .field-grid {{
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(260px, 1fr));
      gap: 10px 12px;
    }}
    .checkbox-grid {{
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
      gap: 8px 10px;
    }}
    .field {{
      border-left: 3px solid #f3f3f3;
    }}
    .field {{
      display: grid;
      grid-template-rows: auto auto auto;
      gap: 4px;
      padding: 10px;
      border: 1px solid var(--border);
      border-radius: 6px;
      background: #fff;
    }}
    .field.checkbox {{
      grid-template-columns: auto 1fr auto;
      grid-template-rows: auto;
      align-items: center;
      gap: 8px;
      background: #f8f9fc;
      border-color: #d5dbe7;
    }}
    .field.checkbox .label {{ font-weight: 600; color: var(--ink); }}
    .label {{ font-size: 13px; color: var(--ink); font-weight: 600; }}
    input[type="text"], input[type="number"] {{
      width: 100%;
      padding: 7px 9px;
      border-radius: 4px;
      border: 1px solid var(--border);
      background: #fdfdff;
      font-size: 14px;
      color: var(--ink);
    }}
    input:focus {{ outline: 2px solid #b5c7e3; }}
    .token {{
      font-size: 12px;
      color: var(--muted);
      font-family: "Consolas", "SFMono-Regular", monospace;
      display: none; /* hide placeholders while keeping them in the DOM */
    }}
    .header-fill {{
      background: var(--header-fill);
      border: 1px solid var(--border);
      border-radius: 6px;
      padding: 6px 10px;
      font-weight: 700;
      color: var(--accent);
      display: inline-block;
      margin-top: 6px;
    }}
    input[type="checkbox"] {{
      width: 18px;
      height: 18px;
    }}
    @media print {{
        .no-print {{ display: none; }}
        .section.collapsed {{ display: none; }}
        body {{ background: white; padding: 0; }}
        .page {{ max-width: 100%; }}
        .section {{ break-inside: avoid; border: 1px solid #ccc; }}
        .field {{ border: none; }}
        input {{ background: transparent; }}
    }}

    @media (max-width: 640px) {{
      body {{ padding: 12px; }}
      .section {{ padding: 12px; }}
      .field-grid,
      .checkbox-grid {{
        grid-template-columns: 1fr;
      }}
      input[type="text"], input[type="number"] {{
        padding: 10px 12px;
        font-size: 15px;
      }}
      .field {{
        gap: 6px;
      }}
      .field.checkbox {{
        grid-template-columns: auto 1fr;
        grid-template-rows: auto auto;
        row-gap: 4px;
      }}
    }}
  </style>
</head>
  <body>
    <div class="page">
      <header>
        <h1>General Order Acknowledgement Form</h1>
      </header>
      <div class="divider"></div>
      {body}
    </div>
    <script>
      document.addEventListener('DOMContentLoaded', function () {
          const sections = document.querySelectorAll('.section');
      
          sections.forEach((section, index) => {
              const header = section.querySelector('.section-header');
              if (header) {
                  // Create content wrapper
                  const content = document.createElement('div');
                  content.className = 'section-content';
      
                  // Move all group elements into the content wrapper
                  const groups = Array.from(section.children).filter(child => child.classList.contains('group'));
                  groups.forEach(group => {
                      content.appendChild(group);
                  });
                  section.appendChild(content);
      
                  // Add toggle icon
                  const icon = document.createElement('span');
                  icon.className = 'toggle-icon';
                  icon.textContent = '+';
                  header.appendChild(icon);

                  // Add click listener
                  header.addEventListener('click', () => {
                      section.classList.toggle('collapsed');
                      header.classList.toggle('active');
                  });
      
                  // Initially collapse all but the first section
                  if (index > 0) {
                      section.classList.add('collapsed');
                  } else {
                      header.classList.add('active');
                  }
              }
          });
      });
    </script>
  </body>
</html>
"""


def main():
    rows = load_rows()
    html_doc = build_html(rows)
    OUTPUT_HTML.write_text(html_doc, encoding="utf-8")
    print(f"Wrote {OUTPUT_HTML} with {len(rows)} fields.")


if __name__ == "__main__":
    main()
